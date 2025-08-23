"""
Gestor para el manejo de archivos Excel con alertas geotécnicas
"""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
import hashlib
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Constantes de validación
VALID_ALERT_TYPES = ['Roja', 'Amarilla', 'Naranja']
VALID_CONDITIONS = [
    'Transgresiva', 
    'Progresiva', 
    'Crítica', 
    'Regresiva', 
    'Transgresiva-Progresiva', 
    'Progresiva-Crítica'
]


class ExcelManager:
    """Gestor para operaciones con Excel"""
    
    def __init__(self, excel_file: str = "data/alertas_geotecnicas.xlsx"):
        self.excel_file = Path(excel_file)
        self.excel_file.parent.mkdir(exist_ok=True, parents=True)
        self._ensure_excel_file()
        # Verificar y actualizar estructura si es necesario
        self.update_excel_structure()
        
    def _ensure_excel_file(self):
        """Asegura que el archivo Excel existe con la estructura correcta"""
        if not self.excel_file.exists():
            self._create_empty_excel()
            
    def _create_empty_excel(self):
        """Crea un archivo Excel vacío con la estructura correcta"""
        columns = [
            "FechaHora", "TipoAlerta", "Condicion", "Ubicacion", "VelocidadMmDia",
            "Respaldo", "Colapso", "FechaHoraColapso", "Evacuacion", 
            "CronologiaAnalisis", "Observaciones", "Usuario", "FechaRegistro", "HojaOrigen"
        ]
        
        df = pd.DataFrame(columns=columns)
        
        # Crear el workbook y worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alertas"
        
        # Agregar headers
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2E2E2", end_color="E2E2E2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        # Ajustar ancho de columnas
        column_widths = {
            'A': 18,  # FechaHora
            'B': 12,  # TipoAlerta
            'C': 15,  # Condicion
            'D': 25,  # Respaldo
            'E': 10,  # Colapso
            'F': 18,  # FechaHoraColapso
            'G': 12,  # Evacuacion
            'H': 30,  # CronologiaAnalisis
            'I': 30,  # Observaciones
            'J': 15,  # Usuario
            'K': 18,  # FechaRegistro
            'L': 15   # HojaOrigen
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        wb.save(self.excel_file)
        
    def _get_row_hash(self, alert_data: Dict) -> str:
        """Genera un hash para detectar duplicados"""
        # Usar FechaHora, TipoAlerta y Observaciones para el hash
        hash_string = f"{alert_data['FechaHora']}{alert_data['TipoAlerta']}{alert_data['Observaciones']}"
        return hashlib.md5(hash_string.encode()).hexdigest()
        
    def load_data(self) -> pd.DataFrame:
        """Carga los datos del archivo Excel filtrando cabeceras y separadores"""
        try:
            df = pd.read_excel(self.excel_file, sheet_name="Alertas")
            
            # Eliminar filas completamente vacías
            df = df.dropna(how='all')
            
            # Filtro mejorado para excluir cabeceras y separadores
            if not df.empty:
                # Una fila válida debe tener:
                # 1. FechaHora válida (no solo fechas 01/01/xxxx 00:00:00)
                # 2. TipoAlerta válida (Roja, Amarilla, Naranja)
                # 3. Condicion válida
                
                mask_valid_rows = pd.Series([True] * len(df))
                
                # Filtrar solo texto de meses en FechaHora (mantener separadores 01/01/yyyy)
                if 'FechaHora' in df.columns:
                    fecha_str = df['FechaHora'].astype(str)
                    # Excluir solo texto de meses en FechaHora (mantener separadores de fecha)
                    mask_texto_mes = fecha_str.str.contains(r'enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre', case=False, na=False)
                    mask_valid_rows &= ~mask_texto_mes
                
                # Verificar si es un separador de fecha (01/01/yyyy 00:00)
                fecha_str = df['FechaHora'].astype(str)
                mask_separador = fecha_str.str.contains(r'01/01/\d{4} 00:00', na=False)
                
                # Para separadores, permitir valores vacíos en TipoAlerta y Condicion
                # Para registros normales, requerir valores válidos
                if 'TipoAlerta' in df.columns:
                    mask_tipo_valido = df['TipoAlerta'].isin(VALID_ALERT_TYPES)
                    # Permitir TipoAlerta vacía solo para separadores
                    mask_tipo_ok = mask_tipo_valido | mask_separador
                    mask_valid_rows &= mask_tipo_ok
                
                # Para Condicion, aplicar la misma lógica
                if 'Condicion' in df.columns:
                    mask_condicion_valida = df['Condicion'].isin(VALID_CONDITIONS)
                    # Permitir Condicion vacía solo para separadores
                    mask_condicion_ok = mask_condicion_valida | mask_separador
                    mask_valid_rows &= mask_condicion_ok
                
                # Aplicar todos los filtros
                df = df[mask_valid_rows]
                
                print(f"Filtrado de datos: {len(df)} registros válidos (incluye separadores) después de excluir cabeceras")
            
            # Actualizar usuarios importados a "admin"
            if 'Usuario' in df.columns:
                users_updated = df['Usuario'].str.contains('Importado_', na=False).sum()
                if users_updated > 0:
                    df.loc[df['Usuario'].str.contains('Importado_', na=False), 'Usuario'] = 'admin'
                    print(f"Actualizados {users_updated} usuarios importados a 'admin'")
                    # Guardar los cambios al archivo
                    self._save_user_updates(df)
            
            # Convertir FechaHora a datetime y extraer año/mes
            if not df.empty and 'FechaHora' in df.columns:
                # Convertir a datetime con formato mixto (maneja múltiples formatos)
                df['FechaHora_dt'] = pd.to_datetime(df['FechaHora'], format='mixed', errors='coerce')
                
                # Mantener todas las filas (incluye separadores con fechas no convertibles)
                # df = df[df['FechaHora_dt'].notna()]  # COMENTADO para incluir separadores
                
                # Extraer año y mes para los filtros (solo para fechas válidas)
                df['Año'] = df['FechaHora_dt'].dt.year
                df['Mes'] = df['FechaHora_dt'].dt.month
                
                # Para separadores (01/01/yyyy 00:00), extraer año manualmente
                fecha_str = df['FechaHora'].astype(str)
                mask_separador = fecha_str.str.contains(r'01/01/\d{4} 00:00', na=False)
                if mask_separador.any():
                    # Extraer año de separadores usando regex
                    separador_years = fecha_str.str.extract(r'01/01/(\d{4}) 00:00')[0]
                    # Crear máscara para separadores sin año
                    mask_sin_año = mask_separador & df['Año'].isna()
                    if mask_sin_año.any():
                        # Asignar año a separadores donde el año es NaN
                        años_extraidos = pd.to_numeric(separador_years[mask_sin_año], errors='coerce')
                        df.loc[mask_sin_año, 'Año'] = años_extraidos
                
                # Ordenar por fecha (colocar fechas nulas al final)
                df = df.sort_values('FechaHora_dt', ascending=True, na_position='last')
                
                # Eliminar columna temporal datetime
                df = df.drop('FechaHora_dt', axis=1)
                
            return df
        except Exception as e:
            print(f"Error cargando datos: {e}")
            return pd.DataFrame()
    
    def _save_user_updates(self, df):
        """Guarda las actualizaciones de usuario al archivo Excel"""
        try:
            self._save_formatted_excel(df)
            print("Actualizaciones de usuario guardadas en Excel")
        except Exception as e:
            print(f"Error guardando actualizaciones de usuario: {e}")
            
    def save_alert(self, alert_data: Dict) -> bool:
        """Guarda una nueva alerta con validación de formato"""
        try:
            # Validar y normalizar formato de fecha
            fecha_hora = alert_data.get('FechaHora', '')
            if fecha_hora:
                # Convertir a formato estándar si no lo está
                try:
                    # Intentar parsear con formato mixto
                    fecha_dt = pd.to_datetime(fecha_hora, format='mixed', errors='coerce')
                    if pd.isna(fecha_dt):
                        print(f"⚠️ Fecha inválida detectada: {fecha_hora}")
                        return False
                    # Convertir a formato estándar YYYY-MM-DD HH:MM:SS
                    alert_data['FechaHora'] = fecha_dt.strftime('%Y-%m-%d %H:%M:%S')
                    print(f"📅 Fecha normalizada: {fecha_hora} → {alert_data['FechaHora']}")
                except Exception as e:
                    print(f"❌ Error validando fecha: {e}")
                    return False
            
            # Validar campos obligatorios
            campos_obligatorios = ['TipoAlerta', 'Condicion', 'Ubicacion', 'VelocidadMmDia']
            for campo in campos_obligatorios:
                if not alert_data.get(campo, '').strip():
                    print(f"❌ Campo obligatorio vacío: {campo}")
                    return False
            
            # Validar velocidad numérica
            try:
                velocidad = str(alert_data.get('VelocidadMmDia', '')).replace(',', '.')
                float(velocidad)
                alert_data['VelocidadMmDia'] = velocidad
            except ValueError:
                print(f"❌ Velocidad inválida: {alert_data.get('VelocidadMmDia')}")
                return False
            
            # Cargar datos existentes
            df = self.load_data()
            
            # Agregar columna HojaOrigen si no existe
            if 'HojaOrigen' not in alert_data:
                alert_data['HojaOrigen'] = 'Manual'
            
            # Verificar duplicados
            if self._is_duplicate(alert_data, df):
                print("⚠️ Alerta duplicada detectada")
                return False
                
            # Agregar nueva fila
            new_row = pd.DataFrame([alert_data])
            df = pd.concat([df, new_row], ignore_index=True)
            
            # Ordenar por fecha - usar formato correcto
            if not df.empty and 'FechaHora' in df.columns:
                # Los datos pueden estar en diferentes formatos, usar format='mixed'
                df['FechaHora_dt'] = pd.to_datetime(df['FechaHora'], format='mixed', errors='coerce')
                df = df.sort_values('FechaHora_dt', ascending=True)
                df = df.drop('FechaHora_dt', axis=1)
                print(f"📊 Datos ordenados por fecha - Total: {len(df)} registros")
            
            # Guardar en Excel con formato
            self._save_formatted_excel(df)
            
            return True
            
        except Exception as e:
            print(f"Error guardando alerta: {e}")
            return False
            
    def _is_duplicate(self, alert_data: Dict, df: pd.DataFrame) -> bool:
        """Verifica si la alerta es duplicada"""
        if df.empty:
            return False
            
        # Comparar por fecha, tipo y observaciones
        duplicates = df[
            (df['FechaHora'] == alert_data['FechaHora']) &
            (df['TipoAlerta'] == alert_data['TipoAlerta']) &
            (df['Observaciones'] == alert_data['Observaciones'])
        ]
        
        return len(duplicates) > 0
        
    def _save_formatted_excel(self, df: pd.DataFrame):
        """Guarda el DataFrame con formato en Excel"""
        # Crear una copia para no modificar el DataFrame original
        df_formatted = df.copy()
        
        # Formatear fechas a dd/mm/yyyy hh:mm
        if 'FechaHora' in df_formatted.columns:
            # Convertir a datetime primero
            df_formatted['FechaHora_dt'] = pd.to_datetime(df_formatted['FechaHora'], format='mixed', errors='coerce')
            
            # Formatear las fechas válidas a dd/mm/yyyy hh:mm
            mask_fechas_validas = df_formatted['FechaHora_dt'].notna()
            df_formatted.loc[mask_fechas_validas, 'FechaHora'] = df_formatted.loc[mask_fechas_validas, 'FechaHora_dt'].dt.strftime('%d/%m/%Y %H:%M')
            
            # Para separadores o fechas especiales, mantener formato original si ya está en dd/mm/yyyy
            fecha_str = df_formatted['FechaHora'].astype(str)
            mask_separadores = fecha_str.str.contains(r'\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}', na=False)
            # Los separadores ya están en formato correcto, solo quitar segundos
            df_formatted.loc[mask_separadores, 'FechaHora'] = fecha_str.loc[mask_separadores].str.replace(r'(\d{2}/\d{2}/\d{4} \d{2}:\d{2}):\d{2}', r'\1', regex=True)
            
            # Eliminar columna temporal
            df_formatted = df_formatted.drop('FechaHora_dt', axis=1)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alertas"
        
        # Definir colores mejorados para tipos de alerta
        alert_colors = {
            'Amarilla': PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),  # Dorado
            'Naranja': PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),   # Naranja oscuro
            'Roja': PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")       # Rojo carmesí
        }
        
        # Agregar datos
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Formatear header
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                else:
                    # Colorear según tipo de alerta (columna B)
                    if c_idx == 2 and value in alert_colors:  # Columna TipoAlerta
                        cell.fill = alert_colors[value]
                        if value == 'Roja':
                            cell.font = Font(color="FFFFFF", bold=True)  # Texto blanco para alerta roja
                        elif value == 'Naranja':
                            cell.font = Font(color="FFFFFF", bold=True)  # Texto blanco para naranja
                        else:
                            cell.font = Font(color="000000", bold=True)  # Texto negro para amarilla
                            
        # Ajustar ancho de columnas
        column_widths = {
            'A': 18, 'B': 12, 'C': 15, 'D': 25, 'E': 10, 'F': 18,
            'G': 12, 'H': 30, 'I': 30, 'J': 15, 'K': 18, 'L': 15
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        wb.save(self.excel_file)
        
    def _fuzzy_match_column(self, column_name: str, target_mappings: dict) -> str:
        """Busca la mejor coincidencia para un nombre de columna usando fuzzy matching"""
        import difflib
        
        # Normalizar el nombre de entrada
        clean_name = str(column_name).strip().upper().replace('\n', ' ').replace('\r', '')
        clean_name = ' '.join(clean_name.split())  # Normalizar espacios
        
        best_match = None
        best_score = 0.0
        
        for original, target in target_mappings.items():
            if target is None:  # Columnas que no usamos
                continue
                
            # Normalizar nombre objetivo
            clean_original = str(original).strip().upper().replace('\n', ' ').replace('\r', '')
            clean_original = ' '.join(clean_original.split())
            
            # Calcular similitud
            score = difflib.SequenceMatcher(None, clean_name, clean_original).ratio()
            
            # Bonus por palabras clave
            keywords = {
                'FECHA': ['FECHA', 'DATE', 'HORA', 'TIME'],
                'TIPO': ['TIPO', 'TYPE', 'ALERTA', 'ALERT'],
                'CONDICION': ['CONDICION', 'CONDITION', 'VELOCIDAD'],
                'RESPALDO': ['RESPALDO', 'RESPALTO', 'BACKUP'],
                'COLAPSO': ['COLAPSO', 'COLLAPSE'],
                'EVACUACION': ['EVACUACION', 'EVACUATION'],
                'CRONOLOGIA': ['CRONOLOGIA', 'ANALISIS', 'ANALYSIS', 'CHRONOLOGY'],
                'OBSERVACIONES': ['OBSERVACIONES', 'OBSERVATIONS', 'COMENTARIOS', 'NOTES']
            }
            
            for key_group, keywords_list in keywords.items():
                if any(kw in clean_name for kw in keywords_list) and any(kw in clean_original for kw in keywords_list):
                    score += 0.3
            
            if score > best_score and score > 0.6:  # Umbral mínimo de similitud
                best_score = score
                best_match = target
        
        return best_match

    def _normalize_columns(self, df, sheet_name=""):
        """Normaliza los nombres de columnas de diferentes años con fuzzy matching"""
        # Mapeo expandido con todas las variaciones posibles
        column_mapping = {
            # Fecha y Hora - Variaciones
            'FECHA Y HORA': 'FechaHora',
            'FECHA Y HORA DE ALERTA': 'FechaHora',
            'FECHA HORA': 'FechaHora',
            'FECHAHORA': 'FechaHora',
            'FECHA': 'FechaHora',
            'HORA': 'FechaHora',
            'DATE': 'FechaHora',
            'DATETIME': 'FechaHora',
            
            # Tipo de Alerta - Variaciones
            'TIPO DE\nALERTA': 'TipoAlerta',
            'TIPO DE ALERTA': 'TipoAlerta',
            'TIPO ALERTA': 'TipoAlerta',
            'TIPOALERTA': 'TipoAlerta',
            'TIPO': 'TipoAlerta',
            'ALERTA': 'TipoAlerta',
            'ALERT TYPE': 'TipoAlerta',
            'TYPE': 'TipoAlerta',
            
            # Condición - Variaciones
            'CONDICIÓN': 'Condicion',
            'CONDICION': 'Condicion',
            'CONDITION': 'Condicion',
            'ESTADO': 'Condicion',
            
            # Ubicación - Variaciones (NUEVA)
            'UBICACIÓN': 'Ubicacion',
            'UBICACION': 'Ubicacion',
            'LOCATION': 'Ubicacion',
            'SECTOR': 'Ubicacion',
            'AREA': 'Ubicacion',
            'ZONE': 'Ubicacion',
            'ZONA': 'Ubicacion',
            'LUGAR': 'Ubicacion',
            'SITIO': 'Ubicacion',
            
            # Velocidad mm/día - Variaciones (NUEVA)
            'VELOCIDAD\nmm/día': 'VelocidadMmDia',
            'VELOCIDAD mm/día': 'VelocidadMmDia',
            'VELOCIDAD MM/DÍA': 'VelocidadMmDia',
            'VELOCIDAD MM/DIA': 'VelocidadMmDia',
            'VELOCIDAD (mm/día)': 'VelocidadMmDia',
            'VELOCIDAD (MM/DÍA)': 'VelocidadMmDia',
            'VELOCIDAD': 'VelocidadMmDia',
            'VELOCITY': 'VelocidadMmDia',
            'SPEED': 'VelocidadMmDia',
            'RATE': 'VelocidadMmDia',
            
            # Respaldo - Variaciones
            'RESPALTO': 'Respaldo',
            'RESPALDO': 'Respaldo',
            'BACKUP': 'Respaldo',
            'ARCHIVO': 'Respaldo',
            'DOCUMENTO': 'Respaldo',
            
            # Colapso - Variaciones
            'COLAPSO': 'Colapso',
            'COLLAPSE': 'Colapso',
            'FALLA': 'Colapso',
            'FAILURE': 'Colapso',
            
            # Fecha Hora Colapso - Variaciones
            'FECHA Y HORA\nCOLAPSO': 'FechaHoraColapso',
            'FECHA Y HORA COLAPSO': 'FechaHoraColapso',
            'FECHA HORA COLAPSO': 'FechaHoraColapso',
            'FECHAHORACOLAPSO': 'FechaHoraColapso',
            'FECHA COLAPSO': 'FechaHoraColapso',
            'HORA COLAPSO': 'FechaHoraColapso',
            
            # Evacuación - Variaciones
            'EVACUACIÓN': 'Evacuacion',
            'EVACUACION': 'Evacuacion',
            'EVACUATION': 'Evacuacion',
            'DESALOJO': 'Evacuacion',
            
            # Cronología/Análisis - Variaciones
            'CRONOLOGÍA O\nANÁLISIS': 'CronologiaAnalisis',
            'CRONOLOGÍA O ANÁLISIS': 'CronologiaAnalisis',
            'CRONOLOGIA O ANALISIS': 'CronologiaAnalisis',
            'CRONOLOGÍA': 'CronologiaAnalisis',
            'CRONOLOGIA': 'CronologiaAnalisis',
            'ANÁLISIS': 'CronologiaAnalisis',
            'ANALISIS': 'CronologiaAnalisis',
            'ANALYSIS': 'CronologiaAnalisis',
            'DESCRIPCIÓN': 'CronologiaAnalisis',
            'DESCRIPCION': 'CronologiaAnalisis',
            
            # Observaciones - Variaciones
            'OBSERVACIONES': 'Observaciones',
            'OBSERVATIONS': 'Observaciones',
            'COMENTARIOS': 'Observaciones',
            'COMMENTS': 'Observaciones',
            'NOTAS': 'Observaciones',
            'NOTES': 'Observaciones',
            'DETALLES': 'Observaciones',
            'DETAILS': 'Observaciones',
            
            # Columnas que ignoramos
            'UGB': None,
            'ID': None,
            'NUM': None,
            'NUMERO': None,
            'NUMBER': None,
        }
        
        # Crear DataFrame normalizado
        normalized_df = pd.DataFrame()
        mapping_log = []
        
        # Primer paso: mapeo directo
        for original_col in df.columns:
            clean_col = str(original_col).strip().replace('\n', '\n').replace('\r', '')
            mapped_col = column_mapping.get(clean_col.upper())
            
            if mapped_col is not None:
                normalized_df[mapped_col] = df[original_col]
                mapping_log.append(f"✓ '{original_col}' → '{mapped_col}' (directo)")
            elif mapped_col is None and clean_col.upper() in column_mapping:
                mapping_log.append(f"⊘ '{original_col}' → ignorada")
        
        # Segundo paso: fuzzy matching para columnas no mapeadas
        unmapped_columns = [col for col in df.columns 
                          if str(col).strip().replace('\n', '\n').upper() not in column_mapping]
        
        for original_col in unmapped_columns:
            fuzzy_match = self._fuzzy_match_column(original_col, column_mapping)
            
            if fuzzy_match and fuzzy_match not in normalized_df.columns:
                normalized_df[fuzzy_match] = df[original_col]
                mapping_log.append(f"≈ '{original_col}' → '{fuzzy_match}' (fuzzy)")
            else:
                mapping_log.append(f"? '{original_col}' → no mapeada")
        
        # Agregar columnas faltantes con valores por defecto
        required_columns = [
            'FechaHora', 'TipoAlerta', 'Condicion', 'Ubicacion', 'VelocidadMmDia',
            'Respaldo', 'Colapso', 'FechaHoraColapso', 'Evacuacion', 
            'CronologiaAnalisis', 'Observaciones', 'Usuario', 'FechaRegistro'
        ]
        
        for col in required_columns:
            if col not in normalized_df.columns:
                if col == 'Usuario':
                    normalized_df[col] = 'admin'  # Usuario por defecto para datos importados
                elif col == 'FechaRegistro':
                    from datetime import datetime
                    normalized_df[col] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                elif col == 'Ubicacion':
                    normalized_df[col] = 'No especificada'  # Valor por defecto para ubicación
                elif col == 'VelocidadMmDia':
                    normalized_df[col] = '0'  # Valor por defecto para velocidad
                else:
                    normalized_df[col] = ''
                mapping_log.append(f"+ '{col}' → creada (vacía)")
        
        # Reordenar columnas en el orden correcto
        column_order = required_columns + ['HojaOrigen']
        final_df = pd.DataFrame()
        
        for col in column_order:
            if col in normalized_df.columns:
                final_df[col] = normalized_df[col]
            elif col == 'HojaOrigen':
                final_df[col] = sheet_name
            else:
                final_df[col] = ''
        
        # Log del mapeo para debugging
        print(f"\n--- MAPEO COLUMNAS HOJA '{sheet_name}' ---")
        for log_entry in mapping_log:
            print(log_entry)
        print(f"Total columnas originales: {len(df.columns)}")
        print(f"Total columnas mapeadas: {len([x for x in mapping_log if '→' in x and '?' not in x])}")
        print(f"Filas procesadas: {len(final_df)}")
        print("--- FIN MAPEO ---\n")
        
        return final_df
    
    def import_excel(self, file_path: str) -> tuple[bool, str]:
        """Importa datos desde otro archivo Excel - Lee todas las hojas (años) con normalización"""
        try:
            # Leer todas las hojas del archivo Excel
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            total_new_records = 0
            total_duplicates = 0
            sheets_processed = 0
            processing_log = []
            
            # Cargar datos existentes una sola vez
            existing_df = self.load_data()
            
            for sheet_name in sheet_names:
                try:
                    # Leer cada hoja
                    raw_df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # Saltar hojas vacías
                    if raw_df.empty:
                        processing_log.append(f"Hoja '{sheet_name}': vacía, omitida")
                        continue
                    
                    # Normalizar columnas
                    normalized_df = self._normalize_columns(raw_df, sheet_name)
                    
                    # Verificar que tenemos datos después de normalizar
                    if normalized_df.empty:
                        processing_log.append(f"Hoja '{sheet_name}': sin datos válidos después de normalización")
                        continue
                    
                    # Verificar columnas críticas con lógica más flexible
                    has_critical_data = False
                    rows_with_data = 0
                    
                    # Contar filas que tienen algún contenido válido
                    for _, row in normalized_df.iterrows():
                        row_has_content = False
                        for col in ['FechaHora', 'TipoAlerta', 'Observaciones', 'CronologiaAnalisis']:
                            if col in normalized_df.columns:
                                value = row[col]
                                if pd.notna(value) and str(value).strip():
                                    row_has_content = True
                                    break
                        if row_has_content:
                            rows_with_data += 1
                    
                    if rows_with_data > 0:
                        has_critical_data = True
                    
                    if not has_critical_data:
                        processing_log.append(f"Hoja '{sheet_name}': sin datos válidos en filas, omitida")
                        continue
                    
                    # Procesar registros de esta hoja
                    sheet_new_records = 0
                    sheet_duplicates = 0
                    
                    for _, row in normalized_df.iterrows():
                        alert_data = row.to_dict()
                        
                        # Limpiar valores NaN/NaT
                        for key, value in alert_data.items():
                            if pd.isna(value):
                                alert_data[key] = ""
                            else:
                                alert_data[key] = str(value).strip()
                        
                        # Validación más flexible - solo requiere que tenga ALGÚN contenido útil
                        has_content = False
                        critical_fields = ['FechaHora', 'TipoAlerta', 'Observaciones', 'CronologiaAnalisis', 'Condicion']
                        
                        for field in critical_fields:
                            if alert_data.get(field) and len(alert_data[field]) > 2:  # Al menos 3 caracteres
                                has_content = True
                                break
                        
                        if not has_content:
                            continue  # Saltar fila sin contenido relevante
                        
                        # Si no tiene fecha, intentar usar valor por defecto
                        if not alert_data.get('FechaHora'):
                            alert_data['FechaHora'] = f"01/01/{sheet_name} 00:00:00" if sheet_name.isdigit() else "01/01/2024 00:00:00"
                        
                        # Si no tiene observaciones, usar cronología si existe
                        if not alert_data.get('Observaciones'):
                            if alert_data.get('CronologiaAnalisis'):
                                alert_data['Observaciones'] = alert_data['CronologiaAnalisis']
                            else:
                                alert_data['Observaciones'] = ""  # Dejar en blanco si no hay datos
                        
                        if not self._is_duplicate(alert_data, existing_df):
                            # Asegurar que todas las columnas necesarias existen
                            for col in existing_df.columns:
                                if col not in alert_data:
                                    alert_data[col] = ""
                                    
                            new_row = pd.DataFrame([alert_data])
                            existing_df = pd.concat([existing_df, new_row], ignore_index=True)
                            sheet_new_records += 1
                        else:
                            sheet_duplicates += 1
                    
                    total_new_records += sheet_new_records
                    total_duplicates += sheet_duplicates
                    sheets_processed += 1
                    
                    processing_log.append(f"Hoja '{sheet_name}': {sheet_new_records} nuevos, {sheet_duplicates} duplicados")
                    
                except Exception as e:
                    processing_log.append(f"Hoja '{sheet_name}': error - {str(e)}")
                    continue
                    
            # Ordenar y guardar si hay registros nuevos
            if total_new_records > 0:
                if not existing_df.empty and 'FechaHora' in existing_df.columns:
                    existing_df['FechaHora_dt'] = pd.to_datetime(
                        existing_df['FechaHora'], format='%d/%m/%Y %H:%M:%S', errors='coerce'
                    )
                    existing_df = existing_df.sort_values('FechaHora_dt', ascending=True)
                    existing_df = existing_df.drop('FechaHora_dt', axis=1)
                    
                self._save_formatted_excel(existing_df)
                
            # Crear mensaje de resultado
            message = f"Importación con normalización completada:\n\n"
            message += f"📊 RESUMEN GENERAL:\n"
            message += f"- Hojas procesadas: {sheets_processed}/{len(sheet_names)}\n"
            message += f"- Total registros nuevos: {total_new_records}\n"
            message += f"- Total duplicados omitidos: {total_duplicates}\n\n"
            message += f"🔄 NORMALIZACIÓN APLICADA:\n"
            message += f"- Mapeo automático de columnas\n"
            message += f"- Conversión a formato estándar\n"
            message += f"- Campos faltantes completados\n\n"
            message += f"📋 DETALLE POR HOJA:\n"
            for log_entry in processing_log:
                message += f"- {log_entry}\n"
            message += f"\n💾 Datos guardados en: {self.excel_file}"
            
            return True, message
            
        except Exception as e:
            return False, f"Error en importación: {str(e)}"
            
    def export_excel(self, file_path: str) -> bool:
        """Exporta los datos a un archivo Excel"""
        try:
            df = self.load_data()
            self._save_formatted_excel_to_path(df, file_path)
            return True
        except Exception as e:
            print(f"Error exportando: {e}")
            return False
            
    def _save_formatted_excel_to_path(self, df: pd.DataFrame, file_path: str):
        """Guarda el DataFrame con formato en un archivo específico"""
        # Crear una copia para no modificar el DataFrame original
        df_formatted = df.copy()
        
        # Formatear fechas a dd/mm/yyyy hh:mm
        if 'FechaHora' in df_formatted.columns:
            # Convertir a datetime primero
            df_formatted['FechaHora_dt'] = pd.to_datetime(df_formatted['FechaHora'], format='mixed', errors='coerce')
            
            # Formatear las fechas válidas a dd/mm/yyyy hh:mm
            mask_fechas_validas = df_formatted['FechaHora_dt'].notna()
            df_formatted.loc[mask_fechas_validas, 'FechaHora'] = df_formatted.loc[mask_fechas_validas, 'FechaHora_dt'].dt.strftime('%d/%m/%Y %H:%M')
            
            # Para separadores o fechas especiales, mantener formato original si ya está en dd/mm/yyyy
            fecha_str = df_formatted['FechaHora'].astype(str)
            mask_separadores = fecha_str.str.contains(r'\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}', na=False)
            # Los separadores ya están en formato correcto, solo quitar segundos
            df_formatted.loc[mask_separadores, 'FechaHora'] = fecha_str.loc[mask_separadores].str.replace(r'(\d{2}/\d{2}/\d{4} \d{2}:\d{2}):\d{2}', r'\1', regex=True)
            
            # Eliminar columna temporal
            df_formatted = df_formatted.drop('FechaHora_dt', axis=1)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alertas"
        
        # Definir colores mejorados para tipos de alerta
        alert_colors = {
            'Amarilla': PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),  # Dorado
            'Naranja': PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),   # Naranja oscuro
            'Roja': PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")       # Rojo carmesí
        }
        
        # Agregar datos
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Formatear header
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                else:
                    # Colorear según tipo de alerta (columna B)
                    if c_idx == 2 and value in alert_colors:  # Columna TipoAlerta
                        cell.fill = alert_colors[value]
                        if value == 'Roja':
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif value == 'Naranja':
                            cell.font = Font(color="FFFFFF", bold=True)
                        else:
                            cell.font = Font(color="000000", bold=True)
                            
        # Ajustar ancho de columnas
        column_widths = {
            'A': 18, 'B': 12, 'C': 15, 'D': 25, 'E': 10, 'F': 18,
            'G': 12, 'H': 30, 'I': 30, 'J': 15, 'K': 18, 'L': 15
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        wb.save(file_path)
        
    def get_statistics(self) -> Dict:
        """Obtiene estadísticas de las alertas"""
        df = self.load_data()
        
        if df.empty:
            return {
                'total_alerts': 0,
                'alert_by_type': {},
                'alert_by_user': {},
                'alert_by_condition': {},
                'recent_alerts': 0
            }
            
        stats = {
            'total_alerts': len(df),
            'alert_by_type': df['TipoAlerta'].value_counts().to_dict() if 'TipoAlerta' in df.columns else {},
            'alert_by_user': df['Usuario'].value_counts().to_dict() if 'Usuario' in df.columns else {},
            'alert_by_condition': df['Condicion'].value_counts().to_dict() if 'Condicion' in df.columns else {},
        }
        
        # Calcular alertas recientes (último mes)
        try:
            if 'FechaHora' in df.columns:
                df['FechaHora_dt'] = pd.to_datetime(df['FechaHora'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
                last_month = datetime.now() - pd.Timedelta(days=30)
                recent = df[df['FechaHora_dt'] >= last_month]
                stats['recent_alerts'] = len(recent)
            else:
                stats['recent_alerts'] = 0
        except:
            stats['recent_alerts'] = 0
            
        return stats
    
    def delete_alerts_by_index(self, indices: List[int]) -> bool:
        """Elimina alertas por sus índices en el DataFrame"""
        try:
            df = self.load_data()
            
            if df.empty:
                return False
            
            # Filtrar índices válidos
            valid_indices = [i for i in indices if 0 <= i < len(df)]
            
            if not valid_indices:
                return False
            
            # Eliminar filas
            df_filtered = df.drop(df.index[valid_indices])
            
            # Guardar datos actualizados
            self._save_formatted_excel(df_filtered)
            
            return True
            
        except Exception as e:
            print(f"Error eliminando alertas: {e}")
            return False

    def update_excel_structure(self) -> bool:
        """Actualiza la estructura del Excel añadiendo columnas faltantes"""
        try:
            print("🔄 Verificando estructura del archivo Excel...")
            
            # Cargar datos existentes
            df = self.load_data()
            
            # Columnas esperadas (orden correcto)
            expected_columns = [
                "FechaHora", "TipoAlerta", "Condicion", "Ubicacion", "VelocidadMmDia",
                "Respaldo", "Colapso", "FechaHoraColapso", "Evacuacion", 
                "CronologiaAnalisis", "Observaciones", "Usuario", "FechaRegistro", "HojaOrigen"
            ]
            
            # Verificar si faltan columnas
            missing_columns = [col for col in expected_columns if col not in df.columns]
            
            if not missing_columns:
                print("✅ La estructura del Excel ya está actualizada.")
                return True
            
            print(f"📝 Agregando columnas faltantes: {missing_columns}")
            
            # Agregar columnas faltantes con valores por defecto
            for col in missing_columns:
                if col == 'Ubicacion':
                    df[col] = 'No especificada'
                elif col == 'VelocidadMmDia':
                    df[col] = '0'
                elif col == 'Usuario':
                    df[col] = 'admin'
                elif col == 'FechaRegistro':
                    from datetime import datetime
                    df[col] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                elif col == 'HojaOrigen':
                    df[col] = 'Alertas'
                else:
                    df[col] = ''
            
            # Reordenar columnas según el orden esperado
            df_reordered = pd.DataFrame()
            for col in expected_columns:
                if col in df.columns:
                    df_reordered[col] = df[col]
            
            # Guardar archivo actualizado
            self._save_formatted_excel(df_reordered)
            
            print(f"✅ Excel actualizado correctamente. Agregadas {len(missing_columns)} columnas.")
            return True
            
        except Exception as e:
            print(f"❌ Error actualizando estructura del Excel: {e}")
            return False
